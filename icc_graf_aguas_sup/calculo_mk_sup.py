import openpyxl
import pymannkendall as mk
from extrair_dados.dados_sup import *


def legendar(aba):
    aba.cell(1, 1).value = 'Poço/Ponto'
    aba.cell(1, 2).value = 'Tendência'
    aba.cell(1, 3).value = 'h'
    aba.cell(1, 4).value = 'p-valor'
    aba.cell(1, 5).value = 'Score Mann-Kendall (s)'


class MannKendallSup:
    def __init__(self, lista_pocos=None):
        if lista_pocos is None:
            lista_pocos = ['ASIC1', 'ASIC10', 'ASIC3/4', 'ASIC2', 'ASIC 17']
        self.__lista = lista_pocos

    @property
    def lista_pocos(self):
        return self.__lista

    @lista_pocos.setter
    def lista_pocos(self, lista_pocos):
        self.__lista = lista_pocos

    # tendência pH - conjuntos

    def plot_todos_basicos(self):
        self.tendencia_ph()
        self.tendencia_acidez()
        self.tendencia_manganes()
        self.tendencia_sulfato()
        self.tendencia_ferro()
        self.tendencia_aluminio()
        self.tendencia_condutividade()
        self.tendencia_od()

    def plot_todos_cargas(self):
        self.tendencia_dam_cargas()
        self.tendencia_acidez_cargas()
        self.tendencia_aluminio_cargas()
        self.tendencia_ferro_cargas()
        self.tendencia_manganes_cargas()
        self.tendencia_sulfato_cargas()

    def tendencia_ph(self):
        # roda com -> tendencia_ph2('PMI-02', 'PMI-04', 'PMI-08', etc)
        # inserir local onde a planilha será criada
        # filepath = ""
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_ph[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1,
                        2).value = f'{mk.original_test(dados_ph[item])[0]}'
            sheet1.cell(a + 1,
                        3).value = f'{mk.original_test(dados_ph[item])[1]}'
            sheet1.cell(a + 1,
                        4).value = f'{mk.original_test(dados_ph[item])[2]}'
            sheet1.cell(a + 1,
                        5).value = f'{mk.original_test(dados_ph[item])[5]}'
        wb1.save('sup_Tendência_pH.xlsx')

    # tendência acidez - conjuntos

    def tendencia_acidez(self):
        # inserir local onde a planilha será criada
        # filepath = ""
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_acidez[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1,
                        2).value = f'{mk.original_test(dados_acidez[item])[0]}'
            sheet1.cell(a + 1,
                        3).value = f'{mk.original_test(dados_acidez[item])[1]}'
            sheet1.cell(a + 1,
                        4).value = f'{mk.original_test(dados_acidez[item])[2]}'
            sheet1.cell(a + 1,
                        5).value = f'{mk.original_test(dados_acidez[item])[5]}'
        wb1.save('sup_Tendência_acidez.xlsx')

    # tendência alumínio - conjuntos

    def tendencia_aluminio(self):
        # inserir local onde a planilha será criada
        # filepath = ""
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_aluminio[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1,
                        2).value = f'{mk.original_test(dados_aluminio[item])[0]}'
            sheet1.cell(a + 1,
                        3).value = f'{mk.original_test(dados_aluminio[item])[1]}'
            sheet1.cell(a + 1,
                        4).value = f'{mk.original_test(dados_aluminio[item])[2]}'
            sheet1.cell(a + 1,
                        5).value = f'{mk.original_test(dados_aluminio[item])[5]}'
        wb1.save('sup_Tendência_aluminio.xlsx')

    # tendência ferro - conjuntos

    def tendencia_ferro(self):
        # inserir local onde a planilha será criada
        # filepath = ""
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_ferro[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1,
                        2).value = f'{mk.original_test(dados_ferro[item])[0]}'
            sheet1.cell(a + 1,
                        3).value = f'{mk.original_test(dados_ferro[item])[1]}'
            sheet1.cell(a + 1,
                        4).value = f'{mk.original_test(dados_ferro[item])[2]}'
            sheet1.cell(a + 1,
                        5).value = f'{mk.original_test(dados_ferro[item])[5]}'
        wb1.save('sup_Tendência_ferro.xlsx')

    # tendência manganes - conjuntos

    def tendencia_manganes(self):
        # inserir local onde a planilha será criada
        # filepath = ""
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_manganes[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1,
                        2).value = f'{mk.original_test(dados_manganes[item])[0]}'
            sheet1.cell(a + 1,
                        3).value = f'{mk.original_test(dados_manganes[item])[1]}'
            sheet1.cell(a + 1,
                        4).value = f'{mk.original_test(dados_manganes[item])[2]}'
            sheet1.cell(a + 1,
                        5).value = f'{mk.original_test(dados_manganes[item])[5]}'
        wb1.save('sup_Tendência_manganes.xlsx')

    # tendência sulfato - conjuntos

    def tendencia_sulfato(self):
        # inserir local onde a planilha será criada
        # filepath = ""
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_sulfato[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1,
                        2).value = f'{mk.original_test(dados_sulfato[item])[0]}'
            sheet1.cell(a + 1,
                        3).value = f'{mk.original_test(dados_sulfato[item])[1]}'
            sheet1.cell(a + 1,
                        4).value = f'{mk.original_test(dados_sulfato[item])[2]}'
            sheet1.cell(a + 1,
                        5).value = f'{mk.original_test(dados_sulfato[item])[5]}'
        wb1.save('sup_Tendência_sulfato.xlsx')

    def tendencia_condutividade(self):
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_condutiv[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1,
                        2).value = f'{mk.original_test(dados_condutiv[item])[0]}'
            sheet1.cell(a + 1,
                        3).value = f'{mk.original_test(dados_condutiv[item])[1]}'
            sheet1.cell(a + 1,
                        4).value = f'{mk.original_test(dados_condutiv[item])[2]}'
            sheet1.cell(a + 1,
                        5).value = f'{mk.original_test(dados_condutiv[item])[5]}'
        wb1.save('sup_Tendência_condutividade.xlsx')

    def tendencia_od(self):
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_od[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1,
                        2).value = f'{mk.original_test(dados_od[item])[0]}'
            sheet1.cell(a + 1,
                        3).value = f'{mk.original_test(dados_od[item])[1]}'
            sheet1.cell(a + 1,
                        4).value = f'{mk.original_test(dados_od[item])[2]}'
            sheet1.cell(a + 1,
                        5).value = f'{mk.original_test(dados_od[item])[5]}'
        wb1.save('sup_Tendência_od.xlsx')

    # -----CARGAS-----#

    def tendencia_acidez_cargas(self):
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_cargasacidez[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1,
                        2).value = f'{mk.original_test(dados_cargasacidez[item])[0]}'
            sheet1.cell(a + 1,
                        3).value = f'{mk.original_test(dados_cargasacidez[item])[1]}'
            sheet1.cell(a + 1,
                        4).value = f'{mk.original_test(dados_cargasacidez[item])[2]}'
            sheet1.cell(a + 1,
                        5).value = f'{mk.original_test(dados_cargasacidez[item])[5]}'
        wb1.save('sup_Tendência_cargas_acidez.xlsx')

    def tendencia_aluminio_cargas(self):
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_cargasaluminio[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1,
                        2).value = f'{mk.original_test(dados_cargasaluminio[item])[0]}'
            sheet1.cell(a + 1,
                        3).value = f'{mk.original_test(dados_cargasaluminio[item])[1]}'
            sheet1.cell(a + 1,
                        4).value = f'{mk.original_test(dados_cargasaluminio[item])[2]}'
            sheet1.cell(a + 1,
                        5).value = f'{mk.original_test(dados_cargasaluminio[item])[5]}'
        wb1.save('sup_Tendência_cargas_aluminio.xlsx')

    def tendencia_ferro_cargas(self):
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_cargasferro[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1,
                        2).value = f'{mk.original_test(dados_cargasferro[item])[0]}'
            sheet1.cell(a + 1,
                        3).value = f'{mk.original_test(dados_cargasferro[item])[1]}'
            sheet1.cell(a + 1,
                        4).value = f'{mk.original_test(dados_cargasferro[item])[2]}'
            sheet1.cell(a + 1,
                        5).value = f'{mk.original_test(dados_cargasferro[item])[5]}'
        wb1.save('sup_Tendência_cargas_ferro.xlsx')

    def tendencia_manganes_cargas(self):
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_cargasmanganes[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1,
                        2).value = f'{mk.original_test(dados_cargasmanganes[item])[0]}'
            sheet1.cell(a + 1,
                        3).value = f'{mk.original_test(dados_cargasmanganes[item])[1]}'
            sheet1.cell(a + 1,
                        4).value = f'{mk.original_test(dados_cargasmanganes[item])[2]}'
            sheet1.cell(a + 1,
                        5).value = f'{mk.original_test(dados_cargasmanganes[item])[5]}'
        wb1.save('sup_Tendência_cargas_manganes.xlsx')

    def tendencia_sulfato_cargas(self):
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_cargassulfato[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1,
                        2).value = f'{mk.original_test(dados_cargassulfato[item])[0]}'
            sheet1.cell(a + 1,
                        3).value = f'{mk.original_test(dados_cargassulfato[item])[1]}'
            sheet1.cell(a + 1,
                        4).value = f'{mk.original_test(dados_cargassulfato[item])[2]}'
            sheet1.cell(a + 1,
                        5).value = f'{mk.original_test(dados_cargassulfato[item])[5]}'
        wb1.save('sup_Tendência_cargas_sulfato.xlsx')

    def tendencia_dam_cargas(self):
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            try:
                mk.original_test(dados_cargasDAM[item])
                sheet1.cell(a + 1, 1).value = f'{item}'
                sheet1.cell(a + 1,
                            2).value = f'{mk.original_test(dados_cargasDAM[item])[0]}'
                sheet1.cell(a + 1,
                            3).value = f'{mk.original_test(dados_cargasDAM[item])[1]}'
                sheet1.cell(a + 1,
                            4).value = f'{mk.original_test(dados_cargasDAM[item])[2]}'
                sheet1.cell(a + 1,
                            5).value = f'{mk.original_test(dados_cargasDAM[item])[5]}'
                wb1.save('sup_Tendência_cargas_DAM.xlsx')
            except KeyError:
                print('erro ao efetuar uma das operações')


    # -----MULTINÍVEIS-----#

    # tendência pH - multiníveis

    def tendencia_mn_ph(self):
        # inserir local onde a planilha será criada
        # filepath = ""
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_ph[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1,
                        2).value = f'{mk.original_test(dados_ph[item])[0]}'
            sheet1.cell(a + 1,
                        3).value = f'{mk.original_test(dados_ph[item])[1]}'
            sheet1.cell(a + 1,
                        4).value = f'{mk.original_test(dados_ph[item])[2]}'
            sheet1.cell(a + 1,
                        5).value = f'{mk.original_test(dados_ph[item])[5]}'
        wb1.save('sup_Tendência_pH_multiníveis.xlsx')

    # tendência acidez - multiníveis

    def tendencia_mn_acidez(self):
        # inserir local onde a planilha será criada
        # filepath = ""
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_acidez[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1,
                        2).value = f'{mk.original_test(dados_acidez[item])[0]}'
            sheet1.cell(a + 1,
                        3).value = f'{mk.original_test(dados_acidez[item])[1]}'
            sheet1.cell(a + 1,
                        4).value = f'{mk.original_test(dados_acidez[item])[2]}'
            sheet1.cell(a + 1,
                        5).value = f'{mk.original_test(dados_acidez[item])[5]}'
        wb1.save('sup_Tendência_acidez_multiníveis.xlsx')

    # tendência aluminio - multiníveis

    def tendencia_mn_aluminio(self):
        # inserir local onde a planilha será criada
        # filepath = ""
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_aluminio[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1,
                        2).value = f'{mk.original_test(dados_aluminio[item])[0]}'
            sheet1.cell(a + 1,
                        3).value = f'{mk.original_test(dados_aluminio[item])[1]}'
            sheet1.cell(a + 1,
                        4).value = f'{mk.original_test(dados_aluminio[item])[2]}'
            sheet1.cell(a + 1,
                        5).value = f'{mk.original_test(dados_aluminio[item])[5]}'
        wb1.save('sup_Tendência_aluminio_multiníveis.xlsx')

    # tendência  ferro - multiníveis

    def tendencia_mn_ferro(self):
        # inserir local onde a planilha será criada
        # filepath = ""
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_ferro[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1,
                        2).value = f'{mk.original_test(dados_ferro[item])[0]}'
            sheet1.cell(a + 1,
                        3).value = f'{mk.original_test(dados_ferro[item])[1]}'
            sheet1.cell(a + 1,
                        4).value = f'{mk.original_test(dados_ferro[item])[2]}'
            sheet1.cell(a + 1,
                        5).value = f'{mk.original_test(dados_ferro[item])[5]}'
        wb1.save('sup_Tendência_ferro_multiníveis.xlsx')
