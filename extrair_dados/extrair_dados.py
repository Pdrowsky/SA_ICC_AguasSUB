import openpyxl
import openpyxl as xl

wb2 = openpyxl.Workbook()
sheet2 = wb2['Sheet']
todos_resultados = []
lista_pontos = []
lista_parametros = []

class ExtrairDados:
    def __init__(self,
                 codigo=str,
                 nome=str,
                 tipo=str,
                 parametro=str,
                 resultado=str,
                 unidade=str
                 ):
        self.codigo = codigo
        self.nome = nome
        self.tipo = tipo
        self.parametro = parametro
        self.resultado = resultado
        self.unidade = unidade

    def criar_objetos(self):
        wb1 = xl.load_workbook(r'C:\Users\Pedro\Desktop\TRABALHANDO\Programação\AguasSub\SA_ICC_AguasSUB\extrair_dados\dados_crus.xlsx')
        sheet1 = wb1['Data']
        for linha in range(2, sheet1.max_row + 1):
            resultado = ExtrairDados(sheet1.cell(linha, 1).value,
                                     # coluna codigo
                                     sheet1.cell(linha, 4).value,  # coluna nome
                                     sheet1.cell(linha, 5).value,  # coluna tipo
                                     sheet1.cell(linha, 13).value,
                                     # coluna parametro
                                     sheet1.cell(linha, 14).value,
                                     # coluna resultado
                                     sheet1.cell(linha, 15).value
                                     # coluna unidade
                                     )
            todos_resultados.append(resultado)
            if resultado.nome not in lista_pontos:
                lista_pontos.append(resultado.nome)
            if resultado.parametro not in lista_parametros:
                lista_parametros.append(resultado.parametro)

    def preparar_parametros(self):
        linha = 2
        for parametro in lista_parametros:
            sheet2.cell(linha, 1).value = parametro
            linha += 1

        coluna = 2
        for ponto in lista_pontos:
            sheet2.cell(1, coluna).value = ponto
            coluna += 1

    def colar_objetos(self):
        coluna = 1
        for col in range(2, len(lista_pontos) + 2):
            coluna += 1
            for resultado in todos_resultados:
                for linha in range(2, sheet2.max_row + 1):
                    if sheet2.cell(linha, 1).value == resultado.parametro and \
                            sheet2.cell(1, coluna).value == resultado.nome:
                        sheet2.cell(linha, coluna).value = resultado.resultado
                    else:
                        pass

    def salvar_planilha(self):
        wb2.save('dados_organizados.xlsx')