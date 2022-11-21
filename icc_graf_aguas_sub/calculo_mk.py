import openpyxl
import pymannkendall as mk
from SA_ICC_AguasSUB.icc_graf_aguas_sub.dados import *


def legendar(aba):
    aba.cell(1, 1).value = 'Poço/Ponto'
    aba.cell(1, 2).value = 'Tendência'
    aba.cell(1, 3).value = 'h'
    aba.cell(1, 4).value = 'p-valor'
    aba.cell(1, 5).value = 'Score Mann-Kendall (s)'

class MannKendall:
    def __init__(self, lista_pocos = []):
        self.__lista = lista_pocos


    # tendência pH - conjuntos

    def tendencia_ph(self):
        # roda com -> tendencia_ph2('PMI-02', 'PMI-04', 'PMI-08', etc)
        # inserir local onde a planilha será criada
        # filepath = ""
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_ph[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1, 2).value = f'{mk.original_test(dados_ph[item])[0]}'
            sheet1.cell(a + 1, 3).value = f'{mk.original_test(dados_ph[item])[1]}'
            sheet1.cell(a + 1, 4).value = f'{mk.original_test(dados_ph[item])[2]}'
            sheet1.cell(a + 1, 5).value = f'{mk.original_test(dados_ph[item])[5]}'
        wb1.save('Tendência_pH')


    # tendência acidez - conjuntos

    def tendencia_acidez(self):
        # inserir local onde a planilha será criada
        # filepath = ""
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_acidez[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1,
                        2).value = f'{mk.original_test(dados_acidez[item])[0]}'
            sheet1.cell(a + 1,
                        3).value = f'{mk.original_test(dados_acidez[item])[1]}'
            sheet1.cell(a + 1,
                        4).value = f'{mk.original_test(dados_acidez[item])[2]}'
            sheet1.cell(a + 1,
                        5).value = f'{mk.original_test(dados_acidez[item])[5]}'
        wb1.save('Tendência_acidez')


    # tendência alumínio - conjuntos

    def tendencia_aluminio(self):
        # inserir local onde a planilha será criada
        # filepath = ""
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_aluminio[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1,
                        2).value = f'{mk.original_test(dados_aluminio[item])[0]}'
            sheet1.cell(a + 1,
                        3).value = f'{mk.original_test(dados_aluminio[item])[1]}'
            sheet1.cell(a + 1,
                        4).value = f'{mk.original_test(dados_aluminio[item])[2]}'
            sheet1.cell(a + 1,
                        5).value = f'{mk.original_test(dados_aluminio[item])[5]}'
        wb1.save('Tendência_aluminio')


    # tendência ferro - conjuntos

    def tendencia_ferro(self):
        # inserir local onde a planilha será criada
        # filepath = ""
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_ferro[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1,
                        2).value = f'{mk.original_test(dados_ferro[item])[0]}'
            sheet1.cell(a + 1,
                        3).value = f'{mk.original_test(dados_ferro[item])[1]}'
            sheet1.cell(a + 1,
                        4).value = f'{mk.original_test(dados_ferro[item])[2]}'
            sheet1.cell(a + 1,
                        5).value = f'{mk.original_test(dados_ferro[item])[5]}'
        wb1.save('Tendência_ferro')


    # tendência manganes - conjuntos

    def tendencia_manganes(self):
        # inserir local onde a planilha será criada
        # filepath = ""
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_manganes[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1,
                        2).value = f'{mk.original_test(dados_manganes[item])[0]}'
            sheet1.cell(a + 1,
                        3).value = f'{mk.original_test(dados_manganes[item])[1]}'
            sheet1.cell(a + 1,
                        4).value = f'{mk.original_test(dados_manganes[item])[2]}'
            sheet1.cell(a + 1,
                        5).value = f'{mk.original_test(dados_manganes[item])[5]}'
        wb1.save('Tendência_manganes')


    # tendência sulfato - conjuntos

    def tendencia_sulfato(self):
        # inserir local onde a planilha será criada
        # filepath = ""
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_sulfato[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1,
                        2).value = f'{mk.original_test(dados_sulfato[item])[0]}'
            sheet1.cell(a + 1,
                        3).value = f'{mk.original_test(dados_sulfato[item])[1]}'
            sheet1.cell(a + 1,
                        4).value = f'{mk.original_test(dados_sulfato[item])[2]}'
            sheet1.cell(a + 1,
                        5).value = f'{mk.original_test(dados_sulfato[item])[5]}'
        wb1.save('Tendência_sulfato')


    # -----MULTINÍVEIS-----#

    # tendência pH - multiníveis

    def tendencia_mn_ph(self):
        # inserir local onde a planilha será criada
        # filepath = ""
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_ph[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1, 2).value = f'{mk.original_test(dados_ph[item])[0]}'
            sheet1.cell(a + 1, 3).value = f'{mk.original_test(dados_ph[item])[1]}'
            sheet1.cell(a + 1, 4).value = f'{mk.original_test(dados_ph[item])[2]}'
            sheet1.cell(a + 1, 5).value = f'{mk.original_test(dados_ph[item])[5]}'
        wb1.save('Tendência_pH_multiníveis')


    # tendência acidez - multiníveis

    def tendencia_mn_acidez(self):
        # inserir local onde a planilha será criada
        # filepath = ""
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_acidez[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1,
                        2).value = f'{mk.original_test(dados_acidez[item])[0]}'
            sheet1.cell(a + 1,
                        3).value = f'{mk.original_test(dados_acidez[item])[1]}'
            sheet1.cell(a + 1,
                        4).value = f'{mk.original_test(dados_acidez[item])[2]}'
            sheet1.cell(a + 1,
                        5).value = f'{mk.original_test(dados_acidez[item])[5]}'
        wb1.save('Tendência_acidez_multiníveis')


    # tendência aluminio - multiníveis

    def tendencia_mn_aluminio(self):
        # inserir local onde a planilha será criada
        # filepath = ""
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_aluminio[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1,
                        2).value = f'{mk.original_test(dados_aluminio[item])[0]}'
            sheet1.cell(a + 1,
                        3).value = f'{mk.original_test(dados_aluminio[item])[1]}'
            sheet1.cell(a + 1,
                        4).value = f'{mk.original_test(dados_aluminio[item])[2]}'
            sheet1.cell(a + 1,
                        5).value = f'{mk.original_test(dados_aluminio[item])[5]}'
        wb1.save('Tendência_aluminio_multiníveis')


    # tendência  ferro - multiníveis

    def tendencia_mn_ferro(self):
        # inserir local onde a planilha será criada
        # filepath = ""
        wb1 = openpyxl.Workbook()
        sheet1 = wb1['Sheet']
        for item in self.__lista:
            a = sheet1.max_row
            legendar(sheet1)
            mk.original_test(dados_ferro[item])
            sheet1.cell(a + 1, 1).value = f'{item}'
            sheet1.cell(a + 1,
                        2).value = f'{mk.original_test(dados_ferro[item])[0]}'
            sheet1.cell(a + 1,
                        3).value = f'{mk.original_test(dados_ferro[item])[1]}'
            sheet1.cell(a + 1,
                        4).value = f'{mk.original_test(dados_ferro[item])[2]}'
            sheet1.cell(a + 1,
                        5).value = f'{mk.original_test(dados_ferro[item])[5]}'
        wb1.save('Tendência_ferro_multiníveis')
